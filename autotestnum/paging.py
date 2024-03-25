from openpyxl import load_workbook
from openpyxl.styles import Font, Side, Alignment, Border
from openpyxl.worksheet.pagebreak import Break


# 添加大标题
def add_big_title(worksheet, title, col_num):
    # 插入新行作为大标题行
    worksheet.insert_rows(1)

    # 合并大标题行的单元格
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_num)
    worksheet.row_dimensions[1].height = 36

    # 设置大标题单元格的内容和格式
    big_title_cell = worksheet.cell(row=1, column=1, value=title)
    big_title_cell.font = Font(name='微软雅黑', size=22, bold=True)
    big_title_cell.border = Border(bottom=Side(style='thin'))
    big_title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 需要重新设置最后一行行高
    worksheet.row_dimensions[worksheet.max_row].height = 22


# 打印设置
def print_format(for_sheet):
    # 设置打印格式
    # 设置打印缩放，确保同一个班可以被防灾一页中
    for_sheet.page_setup.scale = 40  # 设置缩放值为40%
    for_sheet.print_options.horizontalCentered = True  # 水平居中
    # 设置页边距（单位：英寸）
    for_sheet.page_margins.top = 1.5 / 2.54  # 上边距1.5厘米
    for_sheet.page_margins.bottom = 1 / 2.54  # 下边距1厘米
    for_sheet.page_margins.left = 1.5 / 2.54  # 左边距1.5厘米
    for_sheet.page_margins.right = 1.5 / 2.54  # 右边距1.5厘米
    for_sheet.page_margins.header = 0.8 / 2.54
    # 添加打印标题（第一行和第二行）
    for_sheet.print_title_rows = '1:2'


def page_format():
    # 加载工作簿
    workbook = load_workbook("考生去向表/考生去向表.xlsx")
    # 处理“考生去向表”
    sheet = workbook["考生去向表"]
    # 标题行
    columns = ["班级", "姓名", "科目组", "分数", "考号", "考室号", "座位号", "楼层", "教室"]
    # 假设 worksheet 是您的工作表对象
    add_big_title(sheet, "考生去向表", len(columns))
    last_class = None
    # 算上大标题，从第三行开始遍历（min_row=3），避免标题被单独分成一页
    # start=2，设置enumerate的索引起始值为2，不是说跳过第一个元素。目的是保持i与表格行索引对齐
    for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=2):
        current_class = row[0]  # 班级是每行的第一列
        if last_class and current_class != last_class:
            sheet.row_breaks.append(Break(id=i))
        last_class = current_class
    print_format(sheet)

    # 设置页眉，当前页+班级
    sheet.oddHeader.right.text = "&P" + ' C'
    sheet.oddHeader.right.size = 22

    # 处理“考室座次表”
    sheet = workbook["考室座次表"]
    # 添加标题行
    add_big_title(sheet, "考室座次表", len(columns))
    last_room = None
    # start=2，设置enumerate的起始值为2，不是说跳过第一个元素。目的是保持i与表格行索引对齐
    for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=2):
        current_room = row[5]  # 考室号是每行的第六列
        if last_room and current_room != last_room:
            sheet.row_breaks.append(Break(id=i))
        last_room = current_room
    print_format(sheet)
    # 设置页眉，当前页+考室
    sheet.oddHeader.right.text = "&P" + '考室'
    sheet.oddHeader.right.size = 22
    # 保存更改
    workbook.save("考生去向表/考生去向表_打印设置.xlsx")
    print("打印设置完毕，请打开“考生去向表”目录下“考生去向表_打印设置.xlsx”进行查看。")


def one_page_print(for_sheet):
    # 设置打印格式
    # 设置页边距（单位：英寸）
    for_sheet.page_margins.top = 1.5 / 2.54  # 上边距1.5厘米
    for_sheet.page_margins.bottom = 0.6 / 2.54  # 下边距1厘米
    for_sheet.page_margins.left = 1.5 / 2.54  # 左边距1.5厘米
    for_sheet.page_margins.right = 1.5 / 2.54  # 右边距1.5厘米
    for_sheet.page_margins.header = 0.8 / 2.54
    # 打印到一页
    for_sheet.page_setup.fitToPage = 1
    for_sheet.page_setup.fitToHeight = 1
    for_sheet.page_setup.fitToWidth = 1
    # 设置纸张大小为A4
    for_sheet.page_setup.paperSize = for_sheet.PAPERSIZE_A4
    # 水平居中
    for_sheet.print_options.horizontalCentered = True
    # 添加打印标题（第一行和第二行）
    for_sheet.print_title_rows = '1:2'


# 创建一个方法，打开文件，遍历文件中所有表格，对每个表格进行打印格式设置
def page_format_cf(file_path, big_title):
    # 小标题行
    columns = ["年级", "班级", "姓名", "科目组", "考号", "考室号", "座位号", "楼层", "教室"]
    # 加载工作簿
    workbook = load_workbook(file_path)
    # 遍历所有工作表
    for sheet in workbook.worksheets:
        # 添加大标题行
        add_big_title(sheet, big_title, len(columns))
        # 设置打印格式
        one_page_print(sheet)
        # 设置页眉，当前页+班级
        sheet.oddHeader.right.text = sheet.title
        sheet.oddHeader.right.size = 22
    # 保存更改
    workbook.save(file_path)
    print("文件：“" + file_path + "”打印格式调整完成。")

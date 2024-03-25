from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def excel_format(file_path):
    workbook = load_workbook(file_path)
    # 设置边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 针对每个工作表进行格式调整
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        # 设置列宽比例
        col_width_ratios = [4, 5, 5, 5, 7, 4, 4, 6, 6]
        first_col_width = 10  # 第一列宽度

        # 设置列宽
        for idx, ratio in enumerate(col_width_ratios, start=1):
            column_width = first_col_width * (ratio / col_width_ratios[0])
            worksheet.column_dimensions[get_column_letter(idx)].width = column_width

        # 遍历所有行
        for row in worksheet.iter_rows():
            # 设置当前行的行高
            worksheet.row_dimensions[row[0].row].height = 22

            # 设置单元格的字体、边框和对齐方式
            for cell in row:
                # 设置字体
                if cell.row == 1:  # 标题行
                    cell.font = Font(name='微软雅黑', size=13, bold=True)
                else:  # 数据行
                    cell.font = Font(name='微软雅黑', size=12)

                # 设置边框
                cell.border = thin_border

                # 设置对齐
                cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存更改

    workbook.save(file_path)
    print("文件：“" + file_path + "”格式调整完成。")

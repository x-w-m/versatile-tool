import openpyxl
from copy import copy

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def find_class_column_index(sheet):
    """查找班级列的索引"""
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "班级":
            return col[0].column
    return None


def get_classes(sheet, class_col_index):
    """获取所有班级"""
    classes = set()
    for row in sheet.iter_rows(min_row=2, max_col=class_col_index, max_row=sheet.max_row):
        class_name = row[class_col_index - 1].value
        if class_name:
            classes.add(class_name)
    return classes


def apply_cell_format(new_cell, template_cell):
    """应用单元格格式"""
    new_cell.font = copy(template_cell.font)
    new_cell.border = copy(template_cell.border)
    new_cell.fill = copy(template_cell.fill)
    new_cell.number_format = copy(template_cell.number_format)
    new_cell.protection = copy(template_cell.protection)
    new_cell.alignment = copy(template_cell.alignment)


def copy_format_and_data(sheet, new_sheet, class_name, class_col_index, title_row_format, data_row_format,
                         title_row_height, data_row_height):
    """复制格式和数据到新的工作表"""
    new_sheet.row_dimensions[1].height = title_row_height
    for idx, template_cell in enumerate(title_row_format, start=1):
        new_cell = new_sheet.cell(row=1, column=idx, value=template_cell.value)
        apply_cell_format(new_cell, template_cell)

    for row in sheet.iter_rows(min_row=2):
        if row[class_col_index - 1].value == class_name:
            new_row_idx = new_sheet.max_row + 1
            new_sheet.row_dimensions[new_row_idx].height = data_row_height
            for idx, cell in enumerate(row, start=1):
                new_cell = new_sheet.cell(row=new_row_idx, column=idx, value=cell.value)
                apply_cell_format(new_cell, data_row_format[idx - 1])

    print_format(class_name, new_sheet)


# 设置打印格式
def print_format(class_name, new_sheet):
    # 设置打印格式
    new_sheet.page_setup.fitToPage = True
    new_sheet.page_setup.fitToWidth = 1  # 调整为一页宽
    new_sheet.page_setup.fitToHeight = 1
    new_sheet.print_options.horizontalCentered = True  # 水平居中
    # 设置页边距（单位：英寸）
    new_sheet.page_margins.top = 1.5 / 2.54  # 上边距1.5厘米
    new_sheet.page_margins.bottom = 1 / 2.54  # 下边距1厘米
    new_sheet.page_margins.left = 1.5 / 2.54  # 左边距1.5厘米
    new_sheet.page_margins.right = 1.5 / 2.54  # 右边距1.5厘米
    new_sheet.page_margins.header = 0.8 / 2.54
    # 设置页眉
    new_sheet.oddHeader.right.text = class_name + " C"
    new_sheet.oddHeader.right.size = 16  # 字号
    new_sheet.oddHeader.right.font = "宋体,加粗"
    new_sheet.evenHeader.right.text = class_name + " C"
    new_sheet.evenHeader.right.size = 16  # 字号
    new_sheet.evenHeader.right.font = "宋体,加粗"


def split_excel_by_class(file_path):
    """主程序：按班级拆分Excel表格"""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["考生去向表"]

    class_col_index = find_class_column_index(sheet)
    if class_col_index is None:
        raise ValueError("未找到 '班级' 列")

    classes = get_classes(sheet, class_col_index)
    column_widths = {get_column_letter(idx): sheet.column_dimensions[get_column_letter(idx)].width for idx in
                     range(1, sheet.max_column + 1)}
    title_row_format = [copy(cell) for cell in sheet[1]]
    title_row_height = sheet.row_dimensions[1].height
    data_row_format = [copy(cell) for cell in sheet[2]]
    data_row_height = sheet.row_dimensions[2].height

    for class_name in classes:
        new_sheet = workbook.create_sheet(title=str(class_name))
        for letter, width in column_widths.items():
            new_sheet.column_dimensions[letter].width = width
        copy_format_and_data(sheet, new_sheet, class_name, class_col_index, title_row_format, data_row_format,
                             title_row_height, data_row_height)

    new_file_path = file_path.replace(".xlsx", "_按班级拆分.xlsx")
    workbook.save(new_file_path)
    return new_file_path


if __name__ == "__main__":
    file_path = "考生去向表_格式调整.xlsx"
    new_file_path = split_excel_by_class(file_path)
    print(f"文件已保存为: {new_file_path}")

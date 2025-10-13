"""
各排名段人数统计
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# 考试文件和对应的工作表信息
exam_files = [
    {'file': '考试成绩/2323级分科分班成绩.xlsx', 'sheets': ['物理', '历史'], 'exam_name': '分班'},
    {'file': '考试成绩/2024年10月高二月考班级赋分成绩.xlsx', 'sheets': ['物理', '历史'], 'exam_name': '10月月考'},
    {'file': '考试成绩/2024年11月高二期中考试班级赋分成绩.xlsx', 'sheets': ['物理', '历史'], 'exam_name': '期中考试'},
    {'file': '考试成绩/2024年12月高二月考班级赋分成绩.xlsx', 'sheets': ['物理', '历史'], 'exam_name': '12月月考'},
]

# 定义列名
headers = ['班级', '姓名', '考号', '选科', '原始总分', '赋分总分', '班排名', '校排名', '语文', '数学', '外语',
           '选科', '化学原始分', '化学', '生物原始分', '生物', '政治原始分', '政治', '地理原始分', '地理']

# 统计函数
def count_students_in_rank(data, rank_thresholds, rank_col='校排名'):
    results = {}
    for rank in rank_thresholds:
        filtered_data = data[data[rank_col] <= rank]
        group_counts = filtered_data.groupby('班级')[rank_col].count()
        results[f'前{rank}名'] = group_counts
    return pd.DataFrame(results)

# 创建一个空的字典，用于存储所有统计结果
all_statistics = {}

# 遍历考试文件和工作表
for exam in exam_files:
    file = exam['file']
    sheets = exam['sheets']
    exam_name = exam['exam_name']

    for sheet_name in sheets:
        # 读取表格数据
        data = pd.read_excel(file, sheet_name=sheet_name, header=None, skiprows=3)
        data.columns = headers

        # 根据科目设置排名分段
        if sheet_name == '物理':
            rank_thresholds = [50, 100, 200, 500, 800]
        elif sheet_name == '历史':
            rank_thresholds = [30, 60, 120, 300, 480]
        else:
            continue

        # 统计结果
        result = count_students_in_rank(data, rank_thresholds)
        # 将结果存入字典，以方便统一保存
        all_statistics[f'{exam_name}_{sheet_name}'] = result

# 添加阶段人数增减统计
phase_changes = {}
exam_names = [exam['exam_name'] for exam in exam_files]

for i in range(1, len(exam_files)):
    prev_exam = exam_files[i - 1]['exam_name']
    curr_exam = exam_files[i]['exam_name']

    for sheet_name in exam_files[i]['sheets']:
        prev_key = f'{prev_exam}_{sheet_name}'
        curr_key = f'{curr_exam}_{sheet_name}'

        if prev_key in all_statistics and curr_key in all_statistics:
            prev_data = all_statistics[prev_key]
            curr_data = all_statistics[curr_key]

            # 对两个阶段的数据对齐
            prev_data, curr_data = prev_data.align(curr_data, fill_value=0)

            # 计算增减情况
            changes = curr_data - prev_data
            phase_changes[f'{curr_exam}_相较_{prev_exam}_{sheet_name}'] = changes

# 保存所有统计结果到一个 Excel 文件
output_file = '人数统计.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, result in all_statistics.items():
        result.to_excel(writer, sheet_name=sheet_name)

    # 保存增减统计
    for sheet_name, changes in phase_changes.items():
        changes.to_excel(writer, sheet_name=sheet_name)

# 设置表格样式
def set_excel_styles(file_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name='微软雅黑', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    wb.save(file_path)

# 应用样式
set_excel_styles(output_file)

print(f"统计结果已保存到 {output_file}，并设置了样式。")

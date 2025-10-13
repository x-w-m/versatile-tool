"""
各班各科临界生，成绩表表名必须为“物理”，“历史”。
"""
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def analyze_critical_students(grades_file, thresholds_file):
    # 设置 pandas 显示选项
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)

    # 读取成绩表和临界线分数表
    grades_df = pd.ExcelFile(grades_file)
    thresholds_df = pd.read_excel(thresholds_file)

    # 初始化 ExcelWriter
    output_file = thresholds_file.replace("临界分数线", "临界学生名单")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 先添加临界分数线表
        thresholds_df.to_excel(writer, sheet_name='临界分数线', index=False)
        for sheet_name in grades_df.sheet_names:
            # 定义动态科目顺序
            subject_order = ['赋分总分', '语文', '数学', '外语', sheet_name, '化学', '生物', '政治', '地理']

            # 读取每个方向的成绩表
            df = pd.read_excel(grades_df, sheet_name=sheet_name, header=None, skiprows=3)

            # 定义列名
            headers = ['班级', '姓名', '考号', '选科', '原始总分', '赋分总分', '班排名', '校排名', '联考名', '语文',
                       '数学', '外语', sheet_name, '化学原始分', '化学', '生物原始分', '生物', '政治原始分', '政治',
                       '地理原始分', '地理']
            df.columns = headers

            # 根据方向筛选对应的临界项目和临界分数
            thresholds = thresholds_df[thresholds_df['方向'] == sheet_name]

            # 遍历临界项目
            for project_name in thresholds['临界项目'].unique():
                project_thresholds = thresholds[thresholds['临界项目'] == project_name]

                # 初始化单个临界项目的结果
                critical_students = pd.DataFrame()

                # 遍历科目列，按动态科目顺序处理
                for subject in subject_order:
                    if subject == sheet_name:  # 处理物理或历史
                        threshold_column = '物理/历史'
                    else:
                        threshold_column = subject

                    if threshold_column in project_thresholds.columns and subject in df.columns:
                        threshold = project_thresholds[threshold_column].iloc[0]  # 获取对应科目的临界分数

                        # 筛选临界学生
                        mask = df[subject].between(threshold - 15, threshold + 5)
                        subject_critical = df.loc[mask, ['班级', '姓名', '考号', subject]]

                        # 添加临界项目和科目标识
                        subject_critical['临界项目'] = project_name
                        subject_critical['科目'] = subject

                        # 调整列顺序
                        subject_critical = subject_critical[['临界项目', '科目', '班级', '姓名', '考号', subject]]

                        # 合并结果
                        critical_students = pd.concat([critical_students, subject_critical], ignore_index=True)

                # 按班级和科目统计当前项目的临界生人数
                critical_summary = pd.DataFrame()
                if not critical_students.empty:
                    critical_summary = critical_students.groupby(['班级', '科目']).size().unstack(fill_value="")

                # 确保结果按动态科目顺序排序
                if not critical_students.empty:
                    critical_students['科目'] = pd.Categorical(critical_students['科目'], categories=subject_order,
                                                               ordered=True)
                    critical_students = critical_students.sort_values(by=['科目', '班级']).reset_index(drop=True)

                if not critical_summary.empty:
                    critical_summary = critical_summary.reindex(columns=subject_order, fill_value="").reset_index()

                # 保存临界学生名单和统计结果到单独的表
                if not critical_students.empty:
                    critical_students.to_excel(writer, sheet_name=f"{sheet_name}_{project_name}_名单", index=False)

                if not critical_summary.empty:
                    critical_summary.to_excel(writer, sheet_name=f"{sheet_name}_{project_name}_统计", index=False)
    # 应用格式
    apply_excel_format(output_file)


def apply_excel_format(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 确定有效区域
        max_row = ws.max_row
        max_col = ws.max_column
        # 定义样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(name='微软雅黑', size=11, bold=True)
        cell_font = Font(name='微软雅黑', size=11, bold=False)
        alignment = Alignment(horizontal='center', vertical='center')

        # 设置标题行样式
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = alignment
            cell.border = thin_border

        # 设置其他单元格样式
        cell_range = f"A2:{get_column_letter(max_col)}{max_row}"
        for row in ws[cell_range]:
            for cell in row:
                cell.font = cell_font
                cell.alignment = alignment
                cell.border = thin_border

    wb.save(file_path)


if __name__ == "__main__":
    # 成绩表
    grades_file = './班级成绩/2025年03月高二素质检测班级赋分成绩.xlsx'

    # 临界线分数表
    thresholds_file = './临界生/临界分数线/2025年03月高二月考临界分数线.xlsx'

    # 执行分析
    analyze_critical_students(grades_file, thresholds_file)

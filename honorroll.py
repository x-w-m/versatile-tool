# 成绩分析——光荣榜，适用于未分科场景
# 读取Excel表格
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 设置控制台的宽度和每列数据的最大宽度，None表示不限制
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.width', None)

# 读取Excel表格，跳过前三行
df = pd.read_excel('2023年10月高一联考班级成绩.xlsx', header=None, skiprows=3)

# 获取表格第一行、第二行和第三行作为标题
headers = ['班级', '姓名', '考号', '总分', '班排名', '校排名', '语文', '数学', '外语', '物理', '化学', '生物', '政治',
           '历史', '地理']

# 重命名列名
df.columns = headers

# 将考号列设置为索引,不指定也有默认索引，从0递增。
df.set_index('考号', inplace=True)

# 校排名前52名学生
# 按校排名进行排序，然后获取指定校排名学生的信息,默认升序排序，inplace参数默认为False即默认生成新的DataFrame。
df1 = df.sort_values('校排名')
top_52 = df1[df1['校排名'] <= 52][['班级', '姓名', '总分']]
top_52.to_excel('前52.xlsx', index=False, header=['班级', '姓名', '总分'])

# 校排名前1000名学生
top_1000 = df1[df1['校排名'] <= 1000][['班级', '姓名', '总分']]
top_1000.to_excel('前1000.xlsx', index=False, header=['班级', '姓名', '总分'])

# 班级前10
# 按班级和班排名升序排序，并获取班级前10
df2 = df.sort_values(['班级', '班排名'])
top_10 = df2[df2['班排名'] <= 10][['班级', '姓名', '总分', '班排名', '校排名']]
top_10.to_excel('班级前10.xlsx', index=False, header=['班级', '姓名', '分数', '班排名', '校排名'])

# 全校单科王
school_max = pd.DataFrame(columns=['科目', '班级', '姓名', '分数'])
for subject in ['语文', '数学', '外语', '物理', '化学', '生物', '政治', '历史', '地理']:
    subject_max = df[subject].max()
    # 获取最高分对应的学生信息
    temp = df[['班级', '姓名', subject]].loc[df[subject] == subject_max]
    # 添加科目列
    temp.insert(0, '科目', subject)
    # 重命名列名
    temp.columns = ['科目', '班级', '姓名', '分数']
    # 将结果添加到结果 DataFrame 中
    school_max = pd.concat([school_max, temp])

# 各班单科王
# 选择需要的列，包括科目、班级、姓名和分数
school_max = school_max[['科目', '班级', '姓名', '分数']]
# 重置索引，reset_index会将索引转为普通列，需要使用drop=True来删除，不然数据会多出一列。
school_max.reset_index(inplace=True, drop=True)
school_max.to_excel('班级单科王.xlsx', index=False, header=['科目', '班级', '姓名', '分数'])

# 按班级分组，并获取每个班级语文的最高分
class_max = df.groupby('班级')[['语文', '数学', '外语', '物理', '化学', '生物', '政治', '历史', '地理']].max()

# 对每个班级进行聚合操作，按不同科目进行遍历，将每个科目最高分对应的学生姓名使用换行符连接成一个字符串。
bdankewang = pd.DataFrame()
for subject in ['语文', '数学', '外语', '物理', '化学', '生物', '政治', '历史', '地理']:
    temp = class_max.reset_index().merge(df, on=['班级', subject]).loc[:, ['班级', '姓名', subject]].groupby(
        ['班级', subject])[
        '姓名'].agg('\n'.join).reset_index()
    temp.columns = ['班级', '姓名', subject]
    bdankewang = pd.concat([bdankewang, temp], axis=1)

# 将结果保存为新表格
bdankewang.to_excel('最高分.xlsx', index=False)

# 创建一个 ExcelWriter 对象，指定要保存的文件名
writer = pd.ExcelWriter('光荣榜.xlsx')
top_52.to_excel(writer, sheet_name='浪尖起舞', index=False)
top_1000.to_excel(writer, sheet_name="总分优胜", index=False)
top_10.to_excel(writer, sheet_name="班级前10", index=False)
school_max.to_excel(writer, sheet_name="全校单科王", index=False)
bdankewang.to_excel(writer, sheet_name='班级单科王', index=False)
writer.close()

# 为表格设置单元格格式
# 加载Excel文件
wb = load_workbook('光荣榜.xlsx')

xl = pd.ExcelFile('光荣榜.xlsx')
sheets = xl.sheet_names
xl.close()

# 设置字体和文字样式
font = Font(name='微软雅黑', size=12, color='000000', bold=False)

# 设置内容的对齐方式
align = Alignment(horizontal='center', vertical='center')

# 设置单元格边框
border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# 设置单元格填充颜色和样式
# fill = PatternFill(fill_type='solid', fgColor='FFFF00')

# 遍历所有工作表
for sheet in sheets:
    # 遍历所有单元格，判断是否为空，如果不为空则设置格式
    for row in wb[sheet].iter_rows():
        for cell in row:
            if cell.value is not None:  # 判断单元格是否为空
                cell.font = font  # 设置字体
                cell.alignment = align  # 设置对齐方式
                cell.border = border  # 设置边框
                # cell.fill = fill  # 设置填充颜色

# 保存并关闭Excel文件
wb.save('光荣榜.xlsx')
wb.close()
